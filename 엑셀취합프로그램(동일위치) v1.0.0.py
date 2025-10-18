# !pip install openpyxl
import os
import shutil
import openpyxl
from openpyxl.styles import PatternFill

# 병합된 셀인지 확인하는 함수
def is_merged_cell(cell):
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return True
    return False

path = os.getcwd()
ori_folder_path = os.path.join(path, '양식')
ori_file = os.listdir(ori_folder_path)[0]
ok_folder_path = os.path.join(path, '완료')
check_folder_path = os.path.join(path, '검토')
# 취합할 파일들
files_folder_path = os.path.join(path, '취합')
# 폴더가 없으면 생성
for p in [ori_folder_path, ok_folder_path, check_folder_path, files_folder_path]:
    if not os.path.exists(p):
        os.makedirs(p, exist_ok=True)
files_list = os.listdir(files_folder_path)

############################## ori_cell_dict: 양식 정보 딕셔너리 ##############################
ori_path = os.path.join(ori_folder_path, ori_file)
# ori_df = pd.read_excel(ori_path, header=k)

ori_wb = openpyxl.load_workbook(ori_path)
ori_ws = ori_wb.active

# 셀 값을 딕셔너리로 변환
ori_cell_dict = {}
for row in ori_ws.iter_rows(min_row=1, max_row=ori_ws.max_row, min_col=1, max_col=ori_ws.max_column):
    for cell in row:
        ori_cell_dict[(cell.row, cell.column)] = cell.value
# print(ori_cell_dict)

############################## out_cell_dict: 결과 정보 딕셔너리 ##############################
# 값만 None으로 변경
out_cell_dict = {key: None for key in ori_cell_dict}
# print(out_cell_dict)

############################## flag_cell_dict: 플래그 정보 딕셔너리 ##############################
flag_cell_dict = out_cell_dict.copy()
# print(flag_cell_dict)

############################## cell_dict: 취합할 파일 정보 딕셔너리 ##############################
############### 파일별 for문 작업 ###############
for f in files_list:

    try:
        file_path = os.path.join(files_folder_path, f)
        if f[:2] == '(군':
            pass
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # 셀 값을 딕셔너리로 변환
        cell_dict = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell_dict[(cell.row, cell.column)] = cell.value
        # print(cell_dict)

        # ori_cell_dict, cell_dict이 키가 동일하고 값이 다른 경우 out_cell_dict, flag_cell_dict 갱신
        for key in cell_dict:
            if (cell_dict[key] != ori_cell_dict[key]):    # 원본과 값이 다르면
                if (flag_cell_dict[key] is None):         # 덮어쓴 적 없으면
                    if ori_cell_dict[key] is None:
                        out_cell_dict[key] = cell_dict[key]      # 덮어쓰기
                        flag_cell_dict[key] = (1, f)             # flag에 덮어썼다는 의미의 1 삽입 // flag의 값은 None(덮어쓰지않음) 또는 1(덮어씀)
                    else:
                        if (str(ori_cell_dict[key])[:1] != '='):     # 덮어쓴 적 없으면(수식 셀이 아닐 때)
                            out_cell_dict[key] = cell_dict[key]      # 덮어쓰기
                            flag_cell_dict[key] = (1, f)             # flag에 덮어썼다는 의미의 1 삽입 // flag의 값은 None(덮어쓰지않음) 또는 1(덮어씀)
                        else:
                            pass
                else:                                   # 덮어쓴 적 있으면 이 파일 검토 폴더로 이동(shutil.move) 후 다음 파일로 넘어가기
                    break
        else:           # 위 for문이 무사히 break 없이 실행됐다면,
            shutil.move(os.path.join(files_folder_path, f), os.path.join(ok_folder_path, f))
            continue    # 다음 파일로

        # 위 for문이 break에 걸렸다면,
        shutil.move(os.path.join(files_folder_path, f), os.path.join(check_folder_path, f'{f}_{flag_cell_dict[key][1]}'))

    except:
        continue

############################## 변경 셀 값을 적용 ##############################
# 셀 값 적용
# for (row, col), value in out_cell_dict.items():
#     cell = ori_ws.cell(row=row, column=col)
#     if not is_merged_cell(cell):
#         ori_ws.cell(row=row, column=col).value = value

############################## 변경 셀 색을 파랑색으로 표시 ##############################
# 파란색 배경 설정
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# 셀 서식 적용
for (row, col), value in flag_cell_dict.items():
    cell = ori_ws.cell(row=row, column=col)
    if not is_merged_cell(cell) and value is not None:
        ori_ws.cell(row=row, column=col).fill = blue_fill
        ori_ws.cell(row=row, column=col).value = out_cell_dict[(row, col)]

# 수정된 엑셀 파일 저장
ori_wb.save('output.xlsx')

